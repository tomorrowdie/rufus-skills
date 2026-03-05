"""
parse_catalog.py — Amazon FPT (.xlsm) Template Parser for Rufus Skills

Reads an Amazon Flat-File Template (.xlsm) from any raw/ folder and outputs
two normalized CSV files into the sibling data/ folder:
    - fields.csv      — one row per unique field (collapsed numbered variants)
    - valid_values.csv — one row per (field_name, valid_value) pair

Usage:
    python tools/parse_catalog.py --input <path_to_xlsm>
    python tools/parse_catalog.py --input markets/us/catalog/.../raw/POWER_BANK.xlsm

The parser is fully metadata-driven:
    - Row/column positions are read from the Template sheet's Row 1 config
    - Section boundaries are detected automatically
    - No field names are hardcoded — works for ANY Amazon category template

Compatible with: Python 3.9+
Dependencies: openpyxl, pandas (install via: pip install openpyxl pandas)
"""

from __future__ import annotations

import argparse
import csv
import re
import shutil
import sys
from datetime import date
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, unquote

import openpyxl


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Regex to detect numbered field suffixes: bullet_point1, special_features5, etc.
NUMBERED_SUFFIX_RE = re.compile(r"^(.+?)(\d+)$")

# Regex to strip product-type suffix from Valid Values field names
# e.g. "Colour Map - [ powerbank ]" → "Colour Map"
VALID_VALUES_SUFFIX_RE = re.compile(r"\s*-\s*\[.*?\]\s*$")

# Output file names
FIELDS_CSV = "fields.csv"
VALID_VALUES_CSV = "valid_values.csv"


# ---------------------------------------------------------------------------
# 1. Metadata Extraction (Row 1)
# ---------------------------------------------------------------------------

def parse_row1_metadata(ws) -> dict[str, Any]:
    """
    Extract configuration metadata from Row 1 of the Template sheet.

    Amazon FPT templates encode settings in Row 1 as URL-encoded key=value
    pairs. Key settings include:
        - labelRow:               which row has human-readable labels
        - attributeRow:           which row has API field names
        - dataRow:                where seller data entry begins
        - primaryMarketplaceId:   marketplace code (e.g. ATVPDKIKX0DER)
        - contentLanguageTag:     locale tag (e.g. en_US)

    Also extracts section headers that appear as standalone cell values
    in Row 1 (e.g. "Variation", "Discovery", "Fulfillment").

    Returns:
        Dict with parsed config plus a 'sections' key mapping
        section_name → starting column index (0-based).
    """
    meta: dict[str, Any] = {}
    sections: dict[str, int] = {}

    for col_idx, cell in enumerate(ws[1], start=0):
        val = cell.value
        if val is None:
            continue

        val_str = str(val).strip()

        # URL-encoded config block (contains '=' and '&')
        if "=" in val_str and "&" in val_str:
            decoded = unquote(val_str)
            params = parse_qs(decoded, keep_blank_values=True)
            for k, v_list in params.items():
                # parse_qs returns lists; flatten single values
                meta[k] = v_list[0] if len(v_list) == 1 else v_list

        # Simple key=value pairs (e.g. TemplateType=fptcustom)
        elif "=" in val_str and "&" not in val_str:
            parts = val_str.split("=", 1)
            if len(parts) == 2:
                meta[parts[0].strip()] = parts[1].strip()

        # Section headers: short strings without '=' are section markers
        elif len(val_str) < 50 and "=" not in val_str:
            sections[val_str] = col_idx

    meta["sections"] = sections

    # Provide sensible defaults if config keys are missing
    meta.setdefault("labelRow", "2")
    meta.setdefault("attributeRow", "3")
    meta.setdefault("dataRow", "4")

    return meta


# ---------------------------------------------------------------------------
# 2. Template Sheet Field Harvesting
# ---------------------------------------------------------------------------

def harvest_template_fields(ws, meta: dict[str, Any]) -> list[dict[str, Any]]:
    """
    Read the label row and attribute row from the Template sheet and build
    a list of raw field records.

    Each record: {label, field_name, section, column_index}

    Section assignment uses the section markers from Row 1: each field
    belongs to the section whose start column is the largest value <= the
    field's column index.
    """
    label_row_num = int(meta.get("labelRow", 2))
    attr_row_num = int(meta.get("attributeRow", 3))
    sections: dict[str, int] = meta.get("sections", {})

    # Read raw row data
    labels = [cell.value for cell in ws[label_row_num]]
    api_names = [cell.value for cell in ws[attr_row_num]]

    # Build sorted section boundaries for assignment
    # [(col_index, section_name), ...] sorted by col_index
    section_bounds = sorted(
        [(col_idx, name) for name, col_idx in sections.items()],
        key=lambda x: x[0],
    )

    def get_section(col_idx: int) -> str:
        """Find the section a column belongs to based on boundaries."""
        assigned = "Basic"  # default if no section marker precedes this column
        for bound_col, name in section_bounds:
            if bound_col <= col_idx:
                assigned = name
            else:
                break
        return assigned

    fields = []
    max_col = max(len(labels), len(api_names))

    for i in range(max_col):
        label = labels[i] if i < len(labels) else None
        api_name = api_names[i] if i < len(api_names) else None

        # Skip columns with no API field name
        if not api_name or str(api_name).strip() == "":
            continue

        api_name = str(api_name).strip()
        label = str(label).strip() if label else api_name

        fields.append({
            "label": label,
            "field_name": api_name,
            "section": get_section(i),
            "column_index": i,
        })

    return fields


# ---------------------------------------------------------------------------
# 3. Numbered Field Collapsing
# ---------------------------------------------------------------------------

def collapse_numbered_fields(fields: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Collapse repeated numbered fields into a single entry with max_count.

    Example:
        bullet_point1, bullet_point2, ..., bullet_point5
        → one record: field_name='bullet_point', max_count=5

    Fields that don't have numbered variants get max_count=1.

    Strategy:
        - Try to match each field_name against the NUMBERED_SUFFIX_RE pattern
        - Group by base name
        - For groups with >1 member, collapse; use the label/section from
          the first occurrence (lowest number)
    """
    # Track groups: base_name → list of (number, field_record)
    groups: dict[str, list[tuple[int, dict]]] = {}
    standalone: list[dict[str, Any]] = []

    for f in fields:
        match = NUMBERED_SUFFIX_RE.match(f["field_name"])
        if match:
            base = match.group(1)
            num = int(match.group(2))
            groups.setdefault(base, []).append((num, f))
        else:
            standalone.append(f)

    collapsed: list[dict[str, Any]] = []

    # Process standalone fields
    for f in standalone:
        collapsed.append({
            "label": f["label"],
            "field_name": f["field_name"],
            "section": f["section"],
            "max_count": 1,
        })

    # Process grouped (numbered) fields
    for base, members in groups.items():
        members.sort(key=lambda x: x[0])  # sort by number
        max_count = max(num for num, _ in members)
        first = members[0][1]

        # Clean the label: strip the trailing number
        label = re.sub(r"\s*\d+\s*$", "", first["label"]).strip()

        collapsed.append({
            "label": label if label else base,
            "field_name": base,
            "section": first["section"],
            "max_count": max_count,
        })

    # Sort by original column position (use the first field's position for groups)
    field_order = {}
    for f in fields:
        match = NUMBERED_SUFFIX_RE.match(f["field_name"])
        base = match.group(1) if match else f["field_name"]
        if base not in field_order:
            field_order[base] = f["column_index"]

    collapsed.sort(key=lambda f: field_order.get(f["field_name"], 9999))

    return collapsed


# ---------------------------------------------------------------------------
# 4. Data Definitions Enrichment
# ---------------------------------------------------------------------------

def parse_data_definitions(wb) -> dict[str, dict[str, str]]:
    """
    Read the 'Data Definitions' sheet and return a lookup dict keyed by
    API field name.

    Each value: {
        'group': str,
        'definition': str,
        'accepted_values': str,
        'example': str,
        'required': str
    }
    """
    sheet_name = _find_sheet(wb, ["Data Definitions", "data definitions",
                                   "DataDefinitions", "Data_Definitions"])
    if not sheet_name:
        print("[WARN] No 'Data Definitions' sheet found — skipping enrichment.")
        return {}

    ws = wb[sheet_name]
    definitions: dict[str, dict[str, str]] = {}
    current_group = ""

    for row in ws.iter_rows(min_row=3, values_only=True):
        # Column layout: A=group, B=field_name, C=label, D=definition,
        #                E=accepted, F=example, G=required
        if row[0] and str(row[0]).strip():
            current_group = str(row[0]).strip()
            # Clean group: remove long descriptions after ' - '
            if " - " in current_group:
                current_group = current_group.split(" - ")[0].strip()

        field_name = str(row[1]).strip() if row[1] else ""
        if not field_name:
            continue

        definitions[field_name] = {
            "group": current_group,
            "definition": str(row[3]).strip() if row[3] else "",
            "accepted_values": str(row[4]).strip() if row[4] else "",
            "example": str(row[5]).strip() if row[5] else "",
            "required": str(row[6]).strip() if row[6] else "",
        }

    return definitions


# ---------------------------------------------------------------------------
# 5. Valid Values Extraction
# ---------------------------------------------------------------------------

def parse_valid_values(wb) -> dict[str, list[str]]:
    """
    Read the 'Valid Values' sheet and return a dict mapping cleaned field
    names to their list of allowed values.

    Handles:
        - Field names with product-type suffixes: "Colour Map - [ powerbank ]"
          → key = "Colour Map"
        - Variable-width rows (some fields have 2 values, others 900+)
        - Section header rows (skipped)
    """
    sheet_name = _find_sheet(wb, ["Valid Values", "valid values",
                                   "ValidValues", "Valid_Values"])
    if not sheet_name:
        print("[WARN] No 'Valid Values' sheet found — no enumerations extracted.")
        return {}

    ws = wb[sheet_name]
    vv_map: dict[str, list[str]] = {}

    for row in ws.iter_rows(min_row=1, values_only=True):
        # Column B (index 1) holds the field name
        if len(row) < 3 or not row[1]:
            continue

        raw_field = str(row[1]).strip()
        if not raw_field:
            continue

        # Strip product-type suffix
        clean_field = VALID_VALUES_SUFFIX_RE.sub("", raw_field).strip()

        # Collect all non-empty values from column C onward
        values = []
        for cell in row[2:]:
            if cell is not None and str(cell).strip():
                values.append(str(cell).strip())

        if values:
            vv_map[clean_field] = values

    return vv_map


# ---------------------------------------------------------------------------
# 6. Cross-Reference: Match Valid Values to API Field Names
# ---------------------------------------------------------------------------

def map_valid_values_to_api_names(
    vv_by_label: dict[str, list[str]],
    fields: list[dict[str, Any]],
) -> dict[str, list[str]]:
    """
    Valid Values sheet uses human labels (Row 2 style), but our output CSV
    needs API field names. This function builds a label→api_name lookup from
    the fields list and re-keys the valid values dict.

    Falls back to case-insensitive matching if exact match fails.
    """
    # Build label → api_name lookup
    label_to_api: dict[str, str] = {}
    label_to_api_lower: dict[str, str] = {}

    for f in fields:
        label_to_api[f["label"]] = f["field_name"]
        label_to_api_lower[f["label"].lower()] = f["field_name"]

    result: dict[str, list[str]] = {}

    for label, values in vv_by_label.items():
        # Try exact match first
        api_name = label_to_api.get(label)
        if not api_name:
            # Try case-insensitive
            api_name = label_to_api_lower.get(label.lower())
        if not api_name:
            # Last resort: use slugified label as field name
            api_name = re.sub(r"[^a-z0-9]+", "_", label.lower()).strip("_")

        # Handle collapsed numbered fields: strip trailing number
        match = NUMBERED_SUFFIX_RE.match(api_name)
        if match:
            api_name = match.group(1)

        result[api_name] = values

    return result


# ---------------------------------------------------------------------------
# 7. CSV Output with Backup Protection
# ---------------------------------------------------------------------------

def safe_write_csv(
    filepath: Path,
    headers: list[str],
    rows: list[list[Any]],
) -> None:
    """
    Write a CSV file. If the file already exists, back it up with a date
    suffix before overwriting. Creates parent directories as needed.
    """
    filepath.parent.mkdir(parents=True, exist_ok=True)

    if filepath.exists():
        today = date.today().isoformat()
        stem = filepath.stem
        suffix = filepath.suffix
        backup = filepath.parent / f"{stem}_{today}{suffix}"
        # If today's backup already exists, add a counter
        counter = 1
        while backup.exists():
            backup = filepath.parent / f"{stem}_{today}_{counter}{suffix}"
            counter += 1
        shutil.copy2(filepath, backup)
        print(f"[INFO] Backed up existing {filepath.name} → {backup.name}")

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)

    print(f"[OK]   Wrote {filepath} ({len(rows)} rows)")


# ---------------------------------------------------------------------------
# 8. Sheet Finder Helper
# ---------------------------------------------------------------------------

def _find_sheet(wb, candidates: list[str]) -> str | None:
    """Find the first matching sheet name from a list of candidates."""
    available = wb.sheetnames
    for name in candidates:
        if name in available:
            return name
    # Fallback: case-insensitive partial match
    for name in candidates:
        for sheet in available:
            if name.lower() in sheet.lower():
                return sheet
    return None


# ---------------------------------------------------------------------------
# 9. Main Pipeline
# ---------------------------------------------------------------------------

def parse_catalog(input_path: str | Path) -> dict[str, Any]:
    """
    Main entry point: parse an Amazon .xlsm template and produce output CSVs.

    Args:
        input_path: Path to the .xlsm file (must be inside a raw/ folder).

    Returns:
        Summary dict with parse statistics and output paths.
    """
    input_path = Path(input_path).resolve()

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    if not input_path.suffix.lower() in (".xlsm", ".xlsx"):
        raise ValueError(f"Expected .xlsm or .xlsx file, got: {input_path.suffix}")

    # Derive output directory: swap /raw/ for /data/
    raw_dir = input_path.parent
    if raw_dir.name.lower() != "raw":
        raise ValueError(
            f"Input file must be inside a 'raw/' folder. "
            f"Got parent: {raw_dir.name}"
        )
    data_dir = raw_dir.parent / "data"

    print(f"[INFO] Parsing: {input_path.name}")
    print(f"[INFO] Output → {data_dir}/")

    # --- Load Workbook ---
    wb = openpyxl.load_workbook(str(input_path), read_only=True, data_only=True)
    print(f"[INFO] Sheets found: {wb.sheetnames}")

    # --- Find Template sheet ---
    template_name = _find_sheet(wb, ["Template", "template"])
    if not template_name:
        raise ValueError("No 'Template' sheet found in workbook.")
    ws_template = wb[template_name]

    # --- Pass 1: Row 1 metadata ---
    meta = parse_row1_metadata(ws_template)
    marketplace = meta.get("primaryMarketplaceId", "unknown")
    language = meta.get("contentLanguageTag", "unknown")
    print(f"[INFO] Marketplace: {marketplace}  Language: {language}")
    print(f"[INFO] Config → labelRow={meta.get('labelRow')}, "
          f"attributeRow={meta.get('attributeRow')}, "
          f"dataRow={meta.get('dataRow')}")
    if meta.get("sections"):
        print(f"[INFO] Sections: {list(meta['sections'].keys())}")

    # --- Pass 2: Field harvesting ---
    raw_fields = harvest_template_fields(ws_template, meta)
    print(f"[INFO] Raw fields extracted: {len(raw_fields)}")

    # --- Pass 3a: Collapse numbered fields ---
    collapsed_fields = collapse_numbered_fields(raw_fields)
    print(f"[INFO] After collapsing numbered variants: {len(collapsed_fields)} unique fields")

    # --- Pass 3b: Enrich with Data Definitions ---
    definitions = parse_data_definitions(wb)
    if definitions:
        print(f"[INFO] Data Definitions loaded: {len(definitions)} entries")
        for f in collapsed_fields:
            defn = definitions.get(f["field_name"], {})
            if not defn:
                # Try base name for collapsed fields
                match = NUMBERED_SUFFIX_RE.match(f["field_name"])
                if match:
                    defn = definitions.get(match.group(1), {})
            f["required"] = defn.get("required", "")
            f["accepted_values"] = defn.get("accepted_values", "")
            f["example"] = defn.get("example", "")
            f["definition"] = defn.get("definition", "")

    # --- Pass 3c: Valid Values ---
    vv_by_label = parse_valid_values(wb)
    print(f"[INFO] Valid Values fields loaded: {len(vv_by_label)}")

    vv_by_api = map_valid_values_to_api_names(vv_by_label, raw_fields)

    wb.close()

    # --- Write fields.csv ---
    fields_headers = [
        "label", "field_name", "section", "max_count",
        "required", "accepted_values", "example", "definition",
    ]
    fields_rows = [
        [
            f["label"],
            f["field_name"],
            f["section"],
            f["max_count"],
            f.get("required", ""),
            f.get("accepted_values", ""),
            f.get("example", ""),
            f.get("definition", ""),
        ]
        for f in collapsed_fields
    ]

    safe_write_csv(data_dir / FIELDS_CSV, fields_headers, fields_rows)

    # --- Write valid_values.csv ---
    vv_headers = ["field_name", "valid_value"]
    vv_rows = []
    for field_name, values in sorted(vv_by_api.items()):
        for v in values:
            vv_rows.append([field_name, v])

    safe_write_csv(data_dir / VALID_VALUES_CSV, vv_headers, vv_rows)

    # --- Summary ---
    summary = {
        "input_file": str(input_path),
        "marketplace_id": marketplace,
        "language_tag": language,
        "total_raw_fields": len(raw_fields),
        "total_collapsed_fields": len(collapsed_fields),
        "total_valid_value_entries": len(vv_rows),
        "total_valid_value_fields": len(vv_by_api),
        "output_fields_csv": str(data_dir / FIELDS_CSV),
        "output_valid_values_csv": str(data_dir / VALID_VALUES_CSV),
    }

    print("\n" + "=" * 60)
    print("PARSE COMPLETE")
    print("=" * 60)
    for k, v in summary.items():
        print(f"  {k}: {v}")

    return summary


# ---------------------------------------------------------------------------
# CLI Entry Point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Parse Amazon FPT (.xlsm) template → fields.csv + valid_values.csv",
        epilog="Example: python tools/parse_catalog.py --input markets/us/catalog/.../raw/POWER_BANK.xlsm",
    )
    parser.add_argument(
        "--input", "-i",
        required=True,
        help="Path to the .xlsm file (must be inside a raw/ folder)",
    )
    args = parser.parse_args()

    try:
        parse_catalog(args.input)
    except Exception as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
